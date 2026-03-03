from __future__ import annotations

import os
from pathlib import Path

import pandas as pd

import tempfile
import openpyxl
import formulas
from openpyxl import load_workbook

DEFAULT_FILE_PATH = "/content/CF付財務分析表（経営指標あり）_ReadingData_updated.xlsx"
DEFAULT_SHEET_NAME = "CF計算書"
DEFAULT_TITLE = "キャッシュ・フロー計算書"



import tempfile
import openpyxl
import formulas

def _read_excel_values_as_df(file_path: str, sheet_name: str) -> pd.DataFrame:
    """Excelシートを DataFrame 化して返す（数式セルは“計算結果”を値として取得）。

    優先順位:
      1) data_only=True で取得できる「保存済みキャッシュ値」
      2) キャッシュが無い/空のときは、Pythonの式評価エンジン（formulas）で計算して取得

    注意:
      - ②の計算は多少コストがかかります（ただしExcel本体は不要）。
    """
    # まずはキャッシュ値で読む（速い）
    wb_values = load_workbook(file_path, data_only=True, read_only=True)
    if sheet_name not in wb_values.sheetnames:
        raise ValueError(f"シートが見つかりません: {sheet_name}")
    ws_values = wb_values[sheet_name]
    data = [list(r) for r in ws_values.iter_rows(values_only=True)]
    df = pd.DataFrame(data)

    # キャッシュが無い数式が含まれていそうなら formulas で計算して埋める
    # （本スクリプトは先頭〜50行程度しか使わないため、そこだけ判定）
    wb_formula = load_workbook(file_path, data_only=False, read_only=True)
    ws_formula = wb_formula[sheet_name]

    need_calc = False
    max_r = min(ws_formula.max_row or 0, 60)
    max_c = min(ws_formula.max_column or 0, 10)

    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws_formula.cell(row=r, column=c)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                try:
                    v = df.iat[r - 1, c - 1]
                except Exception:
                    v = None
                if v is None or v == "":
                    need_calc = True
                    break
        if need_calc:
            break

    if not need_calc:
        return df

    # --- formulas で計算（シート名の参照を 'シート名'! 形式へ補正してから読む） ---
    # ※ formulas のパーサが非ASCIIシート名の未クオート参照を扱えないケースがあるため
    wb_edit = load_workbook(file_path, data_only=False)
    sheetnames = sorted(wb_edit.sheetnames, key=len, reverse=True)
    for ws in wb_edit.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    new_v = v
                    for s in sheetnames:
                        if f"'{s}'!" in new_v:
                            continue
                        if f"{s}!" in new_v:
                            new_v = new_v.replace(f"{s}!", f"'{s}'!")
                    if new_v != v:
                        cell.value = new_v

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
        tmp_path = tf.name
    try:
        wb_edit.save(tmp_path)

        model = formulas.ExcelModel().loads(tmp_path).finish()
        sol = model.calculate()

        book = os.path.basename(tmp_path)
        # sol のキーは "'[book]Sheet'!A1" 形式
        # DataFrame全体を計算結果で組み立て（最大範囲はシートのmax_row/max_column）
        ws = wb_edit[sheet_name]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        out = [[None for _ in range(max_col)] for _ in range(max_row)]
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                coord = openpyxl.utils.get_column_letter(c) + str(r)
                key = f"'[{book}]{sheet_name}'!{coord}"
                rng = sol.get(key)
                if rng is not None:
                    val = rng.value
                    # 1セルは array([[x]]) なので取り出す
                    try:
                        out[r - 1][c - 1] = val[0][0]
                    except Exception:
                        out[r - 1][c - 1] = val
                else:
                    out[r - 1][c - 1] = ws.cell(row=r, column=c).value

        return pd.DataFrame(out)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass

def build_html(file_path: str, sheet_name: str = DEFAULT_SHEET_NAME, title: str = DEFAULT_TITLE) -> str:
    """
    Excelの指定シート（B列・C列）からキャッシュフロー計算書のHTMLを生成する。
    もともとのColab表示用スクリプトを、サーバ実行（HTMLファイル出力）でも使えるように最低限拡張。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excelファイルが見つかりません: {file_path}")

    df_raw = _read_excel_values_as_df(file_path, sheet_name)

    # B5, C6 をメタ情報として取得（元スクリプト踏襲）
    header_info_period = df_raw.iloc[4, 1]
    header_info_unit = df_raw.iloc[5, 2]

    # 7行目(index 6) 〜 51行目(index 50)、B列(1)とC列(2)
    df = df_raw.iloc[6:51, [1, 2]].fillna("")

    custom_css = """
    <style>
        .report-container { font-family: "Meiryo", sans-serif; color: #000; }
        .report-title { font-size: 18px; font-weight: bold; margin: 20px 0 5px 0; border-left: 5px solid #000; padding-left: 10px; }
        .report-meta { font-size: 13px; margin-bottom: 5px; display: flex; justify-content: space-between; width: 580px; }
        .excel-table {
            border-collapse: collapse; width: fit-content; display: inline-table; font-size: 13px;
            table-layout: fixed; outline: 2px solid #000; box-shadow: 0 0 0 2px #000; background-color: white;
        }
        .excel-table td {
            border-left: 1px solid #999; border-bottom: 1px solid #999;
            padding: 4px 10px; overflow: hidden; white-space: nowrap; box-sizing: border-box; color: #000;
        }
        .excel-table tr td:last-child { border-right: 1px solid #999; }
        .col-subject { width: 400px; text-align: left; }
        .col-amt { width: 180px; text-align: right; font-family: "Consolas", monospace; }
        .header-row td { background-color: #004080 !important; color: #ffffff !important; font-weight: bold; }
        .total-row { background-color: #e6f3ff !important; font-weight: bold; }
        .grand-total { background-color: #d9ead3 !important; font-weight: bold; }
    </style>
    """

    html_output = '<div class="report-container">'
    html_output += f'<div class="report-title">{title}</div>'
    html_output += f'<div class="report-meta"><span>{header_info_period}</span><span>{header_info_unit}</span></div>'
    html_output += '<table class="excel-table"><tbody>'

    for i, row in df.iterrows():
        subject = str(row.iloc[0]).strip()
        amt = row.iloc[1]

        # 空行はスキップ（元スクリプト踏襲）
        if not subject and (amt == "" or amt is None):
            continue

        row_class = ""
        if "キャッシュ・フロー" in subject and i < 20:
            row_class = "header-row"
        elif "計" in subject:
            row_class = "total-row"
        elif "現金及び現金同等物" in subject:
            row_class = "grand-total"

        # 金額の3桁区切り（元スクリプト踏襲）
        display_amt = amt
        if isinstance(amt, (int, float)) and amt != "":
            try:
                display_amt = "{:,}".format(int(amt))
            except Exception:
                display_amt = amt

        html_output += f'<tr class="{row_class}"><td class="col-subject">{row.iloc[0]}</td>'
        html_output += f'<td class="col-amt">{display_amt}</td></tr>'

    html_output += "</tbody></table></div>"

    # 単体HTMLとして返す（CSS込み）
    return '<!doctype html><html><head><meta charset="utf-8">' + custom_css + "</head><body>" + html_output + "</body></html>"


def main() -> None:
    # 互換性のため、未指定なら従来の /content パスを使う
    file_path = os.environ.get("INPUT_XLSX", DEFAULT_FILE_PATH)
    sheet = os.environ.get("SHEET_NAME", DEFAULT_SHEET_NAME)
    title = os.environ.get("TITLE", DEFAULT_TITLE)

    out_html = os.environ.get("OUTPUT_HTML", "")
    html = build_html(file_path=file_path, sheet_name=sheet, title=title)

    if out_html:
        Path(out_html).write_text(html, encoding="utf-8")
        print(out_html)
    else:
        # Colab等でそのまま実行した場合でも最低限確認できるようstdoutへ
        print(html)


if __name__ == "__main__":
    main()
